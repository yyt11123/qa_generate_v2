"""把中间 jsonl 聚合成 xlsx。单行表头，不合并单元格，每行都填分类。

【输出文件命名规约 (v3 起生效)】
本模块默认就带样式输出（_apply_styles 已合并到主流程），
所以不再使用 _styled 后缀。正式定稿文件统一叫 qa_test_full_v<N>.xlsx，
例如 qa_test_full_v3.xlsx；不要再写 qa_test_full_v3_styled.xlsx 这种重复命名。
dry-run 输出可以叫 qa_test_v<N>_dryrun.xlsx 区分。

【page 列】
xlsx 格式底层不区分"空字符串"和"空 cell"，pandas 默认读会把空 cell 显示为 NaN。
用 Excel UI 打开时显示为空白单元格。如需用 pandas 读出来是 ''，传 keep_default_na=False。
带页码时本模块强制写成字符串 + number_format='@'，确保 "17-18" 不会被识别成减法。
"""

import json
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

COLUMNS = [
    "分类",
    "question(咨询问题)",
    "answers(预期回答)",
    "supporting facts(支撑信息)",
    "document(文件名称)",
    "page(所在页码)",
    "text/img/table(支撑文本/图片/表格)",
    "chunk_id",
]

# 内部存繁体 (与 prompt/CATEGORIES 一致)，写 xlsx 时按此映射转成简体显示。
# 不用通用繁简转换库 (如 opencc)，只硬编码这 8 项分类名，避免全文转换引入新风险。
CATEGORY_DISPLAY_MAP = {
    "案例": "案例",
    "產品": "产品",
    "投保規則": "投保规则",
    "健康核保": "健康核保",
    "財務核保": "财务核保",
    "繳費": "缴费",
    "行政規則": "行政规则",
    "一般查詢": "一般查询",
}

COLUMN_WIDTHS = {
    "分类": 10,
    "question(咨询问题)": 40,
    "answers(预期回答)": 60,
    "supporting facts(支撑信息)": 60,
    "document(文件名称)": 35,
    "page(所在页码)": 8,
    "text/img/table(支撑文本/图片/表格)": 12,
    "chunk_id": 12,
}


def _format_page(page_start, page_end) -> str:
    """page_start == page_end → '15'; 跨页 → '17-18'; 缺失/None → ''.
    强制返回字符串，避免 Excel 把数字当日期或公式。"""
    if page_start is None or page_end is None:
        return ""
    if page_start == page_end:
        return str(page_start)
    return f"{page_start}-{page_end}"


def export_to_xlsx(qa_jsonl_path: str | Path, output_xlsx: str | Path) -> int:
    qa_jsonl_path = Path(qa_jsonl_path)
    output_xlsx = Path(output_xlsx)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    if not qa_jsonl_path.exists():
        logger.warning("no qa jsonl found at %s, writing empty xlsx", qa_jsonl_path)
        pd.DataFrame(columns=COLUMNS).to_excel(output_xlsx, index=False)
        return 0

    rows: list[dict] = []
    with qa_jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            qa = json.loads(line)
            cat_raw = qa.get("category") or ""
            cat_display = CATEGORY_DISPLAY_MAP.get(cat_raw, cat_raw)
            if cat_raw and cat_raw not in CATEGORY_DISPLAY_MAP:
                logger.warning(
                    "category %r not in CATEGORY_DISPLAY_MAP, writing as-is (chunk_id=%s)",
                    cat_raw,
                    qa.get("chunk_id"),
                )
            rows.append(
                {
                    "分类": cat_display,
                    "question(咨询问题)": qa.get("question") or "",
                    "answers(预期回答)": qa.get("answer") or "",
                    "supporting facts(支撑信息)": qa.get("supporting_facts") or "",
                    "document(文件名称)": qa.get("source") or "",
                    "page(所在页码)": _format_page(
                        qa.get("page_start"), qa.get("page_end")
                    ),
                    "text/img/table(支撑文本/图片/表格)": qa.get("type") or "",
                    "chunk_id": qa.get("chunk_id") or "",
                }
            )

    df = pd.DataFrame(rows, columns=COLUMNS)
    df = df.fillna("")
    df["page(所在页码)"] = df["page(所在页码)"].astype(str)
    df.to_excel(output_xlsx, index=False, engine="openpyxl")
    _force_page_text_type(output_xlsx)
    _apply_styles(output_xlsx)
    logger.info("wrote %d rows to %s", len(df), output_xlsx)
    return len(df)


def _force_page_text_type(xlsx_path: Path) -> None:
    """把 page 列每个 cell 的 number_format 设成 '@'（文本格式），
    确保像 '15' 或 '17-18' 不会被 Excel 解释成数字、日期或公式。"""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if "page(所在页码)" not in headers:
        return
    page_col_idx = headers.index("page(所在页码)") + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=page_col_idx)
        cell.number_format = "@"
        if cell.value is not None and not isinstance(cell.value, str):
            cell.value = str(cell.value)
    wb.save(xlsx_path)


def _apply_styles(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    for idx, h in enumerate(headers, start=1):
        width = COLUMN_WIDTHS.get(h)
        if width is not None:
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[1].height = 30

    body_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = body_alignment

    ws.freeze_panes = "A2"
    wb.save(xlsx_path)
