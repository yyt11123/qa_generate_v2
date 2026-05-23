"""把 output/per_file/*.xlsx 合并成 output/qa_test_all.xlsx。

【sheet 结构】
- "QA Data"：主 sheet，纯 QA 数据。保持单文件 xlsx 的 8 列原样，前面追加一列
  `source_file(来源文件)`，内容是 per_file 文件的 stem（不带 .xlsx）。
- "Stats"：独立 sheet，存放总计 + 每文件行数。下游评测脚本只读 "QA Data"
  即可，不会被统计行污染。

【样式】
列宽 / 表头加粗 / 冻结 A2 / 自动换行，复用 exporter._apply_styles 的风格。
不用直接 import _apply_styles（私有函数），这里独立实现一份精简版，
避免和单文件 exporter 耦合。
"""
from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

SOURCE_FILE_COL = "source_file(来源文件)"
QA_SHEET = "QA Data"
STATS_SHEET = "Stats"

COLUMN_WIDTHS = {
    SOURCE_FILE_COL: 32,
    "分类": 10,
    "question(咨询问题)": 40,
    "answers(预期回答)": 60,
    "supporting facts(支撑信息)": 60,
    "document(文件名称)": 35,
    "page(所在页码)": 8,
    "text/img/table(支撑文本/图片/表格)": 12,
    "chunk_id": 12,
}


def aggregate(per_file_dir: str | Path, output_xlsx: str | Path) -> int:
    """合并 per_file_dir 下所有 *.xlsx 到 output_xlsx。

    主 sheet "QA Data" 仅包含 QA 行；统计写到独立 sheet "Stats"。返回 QA 行数。
    """
    per_file_dir = Path(per_file_dir)
    output_xlsx = Path(output_xlsx)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    files = sorted(per_file_dir.glob("*.xlsx"))
    if not files:
        logger.warning("no per_file xlsx under %s, writing empty aggregate", per_file_dir)
        with pd.ExcelWriter(output_xlsx, engine="openpyxl") as w:
            pd.DataFrame(columns=[SOURCE_FILE_COL]).to_excel(w, sheet_name=QA_SHEET, index=False)
            pd.DataFrame(columns=["item", "value"]).to_excel(w, sheet_name=STATS_SHEET, index=False)
        return 0

    frames: list[pd.DataFrame] = []
    per_file_counts: list[tuple[str, int]] = []
    for fp in files:
        df = pd.read_excel(fp, dtype=str, keep_default_na=False, engine="openpyxl")
        df.insert(0, SOURCE_FILE_COL, fp.stem)
        frames.append(df)
        per_file_counts.append((fp.stem, len(df)))
        logger.info("loaded %s (%d rows)", fp.name, len(df))

    merged = pd.concat(frames, ignore_index=True)
    n_rows = len(merged)

    stats_df = _build_stats_df(n_rows, per_file_counts)
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as w:
        merged.to_excel(w, sheet_name=QA_SHEET, index=False)
        stats_df.to_excel(w, sheet_name=STATS_SHEET, index=False)
    _apply_styles(output_xlsx)
    logger.info("merged %d files → %s (%d rows)", len(files), output_xlsx, n_rows)
    return n_rows


def _build_stats_df(total_rows: int, per_file_counts: list[tuple[str, int]]) -> pd.DataFrame:
    rows = [
        ("合计", f"{total_rows} 行 / {len(per_file_counts)} 文件"),
    ]
    rows.extend((name, f"{cnt} 行") for name, cnt in per_file_counts)
    return pd.DataFrame(rows, columns=["item", "value"])


def _apply_styles(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    if QA_SHEET in wb.sheetnames:
        _style_qa_sheet(wb[QA_SHEET])
    if STATS_SHEET in wb.sheetnames:
        _style_stats_sheet(wb[STATS_SHEET])
    wb.save(xlsx_path)


def _style_qa_sheet(ws) -> None:
    headers = [c.value for c in ws[1]]

    for idx, h in enumerate(headers, start=1):
        width = COLUMN_WIDTHS.get(h)
        if width is not None:
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[1].height = 30

    body_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = body_alignment

    if "page(所在页码)" in headers:
        page_col_idx = headers.index("page(所在页码)") + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=page_col_idx)
            cell.number_format = "@"
            if cell.value is not None and not isinstance(cell.value, str):
                cell.value = str(cell.value)

    ws.freeze_panes = "A2"


def _style_stats_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 30
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    import argparse

    from config import BATCH_AGGREGATED_XLSX, BATCH_OUTPUT_DIR

    p = argparse.ArgumentParser()
    p.add_argument("--per-file-dir", default=BATCH_OUTPUT_DIR)
    p.add_argument("--output", default=BATCH_AGGREGATED_XLSX)
    args = p.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    n = aggregate(args.per_file_dir, args.output)
    print(f"aggregated {n} rows → {args.output}")
