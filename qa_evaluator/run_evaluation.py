"""评测入口。纯本地脚本，零 API 调用。

Usage:
    python -m qa_evaluator.run_evaluation
    python -m qa_evaluator.run_evaluation --input output/qa_test_full_v3.xlsx
    python -m qa_evaluator.run_evaluation --regen-report  # 仅重生成 report，跳过指标 / 抽样

输出（默认）：
    output/qa_metrics_machine.json
    output/qa_sample_for_annotation.xlsx
    output/qa_evaluation_report.md
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from qa_evaluator.metrics_machine import run_all
from qa_evaluator.report_generator import render_report
from qa_evaluator.sampler import stratified_sample

logger = logging.getLogger(__name__)

DEFAULT_INPUT_XLSX = "output/qa_test_full_v3.xlsx"
DEFAULT_CHUNKS_JSONL = "inputs/chunks_v3.jsonl"
DEFAULT_METRICS_JSON = "output/qa_metrics_machine.json"
DEFAULT_SAMPLE_XLSX = "output/qa_sample_for_annotation.xlsx"
DEFAULT_REPORT_MD = "output/qa_evaluation_report.md"

CATEGORY_DISPLAY_TRAD_TO_SIMP = {
    "案例": "案例", "產品": "产品", "投保規則": "投保规则", "健康核保": "健康核保",
    "財務核保": "财务核保", "繳費": "缴费", "行政規則": "行政规则", "一般查詢": "一般查询",
}

XLSX_COLUMN_TO_KEY = {
    "分类": "category_simp",
    "question(咨询问题)": "question",
    "answers(预期回答)": "answer",
    "supporting facts(支撑信息)": "supporting_facts",
    "document(文件名称)": "source",
    "page(所在页码)": "page",
    "text/img/table(支撑文本/图片/表格)": "type",
    "chunk_id": "chunk_id",
}

# 抽样 xlsx 输出列：原 8 列 + 4 个标注维度 + 备注
ANNOTATION_COLUMNS = [
    "分类",
    "question(咨询问题)",
    "answers(预期回答)",
    "supporting facts(支撑信息)",
    "document(文件名称)",
    "page(所在页码)",
    "text/img/table(支撑文本/图片/表格)",
    "chunk_id",
    "B1 真实性",
    "B2 正确性",
    "B3 完整性",
    "B4 分类合理性",
    "备注",
]
COLUMN_WIDTHS = {
    "分类": 10,
    "question(咨询问题)": 40,
    "answers(预期回答)": 60,
    "supporting facts(支撑信息)": 60,
    "document(文件名称)": 35,
    "page(所在页码)": 8,
    "text/img/table(支撑文本/图片/表格)": 12,
    "chunk_id": 12,
    "B1 真实性": 10,
    "B2 正确性": 10,
    "B3 完整性": 10,
    "B4 分类合理性": 14,
    "备注": 30,
}


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )


def load_qa_xlsx(path: str | Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for h, v in zip(headers, raw):
            key = XLSX_COLUMN_TO_KEY.get(h, h)
            d[key] = v if v is not None else ""
        d["category"] = d.get("category_simp")
        rows.append(d)
    return rows


def load_chunks(path: str | Path) -> tuple[dict[str, dict], int]:
    chunks_by_id: dict[str, dict] = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            chunks_by_id[obj["chunk_id"]] = obj
    return chunks_by_id, len(chunks_by_id)


def write_metrics_json(metrics: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    serializable = {
        name: {
            "name": m["name"],
            "value": m["value"],
            "threshold": m["threshold"],
            "op": m["op"],
            "severity": m["severity"],
            "passed": m["passed"],
            "violation_count": len(m["violations"]),
            "violations": m["violations"],
        }
        for name, m in metrics.items()
    }
    path.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")


def write_sample_xlsx(sample: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(ANNOTATION_COLUMNS)

    for r in sample:
        ws.append([
            r.get("category_simp") or r.get("category") or "",
            r.get("question") or "",
            r.get("answer") or "",
            r.get("supporting_facts") or "",
            r.get("source") or "",
            r.get("page") or "",
            r.get("type") or "",
            r.get("chunk_id") or "",
            "", "", "", "",   # B1 - B4
            "",                # 备注
        ])

    # styles：沿用 exporter 的样式逻辑
    for idx, h in enumerate(ANNOTATION_COLUMNS, start=1):
        width = COLUMN_WIDTHS.get(h)
        if width is not None:
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[1].height = 30

    body_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = body_alignment

    # page 列文本格式
    page_col = ANNOTATION_COLUMNS.index("page(所在页码)") + 1
    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=page_col)
        c.number_format = "@"
        if c.value is not None and not isinstance(c.value, str):
            c.value = str(c.value)

    ws.freeze_panes = "A2"
    wb.save(path)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="QA 测试集质量评测（纯本地，零 API）")
    p.add_argument("--input", default=DEFAULT_INPUT_XLSX, help="待评测的 QA xlsx 路径")
    p.add_argument("--chunks", default=DEFAULT_CHUNKS_JSONL, help="原 chunks jsonl 路径")
    p.add_argument("--metrics-json", default=DEFAULT_METRICS_JSON)
    p.add_argument("--sample-xlsx", default=DEFAULT_SAMPLE_XLSX)
    p.add_argument("--report-md", default=DEFAULT_REPORT_MD)
    p.add_argument("--regen-report", action="store_true",
                   help="跳过指标计算 / 抽样，仅根据现有 metrics-json + sample-xlsx 重生成报告（用于人工标注完成后回填）")
    p.add_argument("--seed", type=int, default=42)
    return p.parse_args()


def run() -> int:
    setup_logging()
    args = parse_args()

    metrics_path = Path(args.metrics_json)
    sample_path = Path(args.sample_xlsx)
    report_path = Path(args.report_md)

    rows = load_qa_xlsx(args.input)
    chunks_by_id, total_chunks = load_chunks(args.chunks)
    logger.info("loaded %d QA rows from %s; %d chunks from %s", len(rows), args.input, total_chunks, args.chunks)

    if args.regen_report:
        if not metrics_path.exists():
            logger.error("regen-report 需要先有 %s", metrics_path)
            return 2
        with metrics_path.open(encoding="utf-8") as f:
            obj = json.load(f)
        # 反序列化时 violations 已含完整明细
        metrics = obj
    else:
        metrics = run_all(rows, chunks_by_id, total_chunks)
        write_metrics_json(metrics, metrics_path)
        logger.info("metrics written: %s", metrics_path)

        sample = stratified_sample(rows, random_state=args.seed)
        write_sample_xlsx(sample, sample_path)
        logger.info("sample xlsx written: %s (%d rows)", sample_path, len(sample))

    annotation_xlsx = sample_path if sample_path.exists() else None
    render_report(
        metrics=metrics,
        total_qa=len(rows),
        total_chunks=total_chunks,
        annotation_xlsx=annotation_xlsx,
        output_md=report_path,
        all_qa_rows=rows,
    )
    logger.info("report written: %s", report_path)

    print()
    print("=== 机器指标摘要 ===")
    for name, m in metrics.items():
        flag = "OK" if m["passed"] else ("FAIL" if m["severity"] == "failure" else "WARN")
        print(f"  [{flag:>4}] {name:<28} value={m['value']:<8} {m['op']} {m['threshold']}")
    print()
    print(f"输出:")
    print(f"  metrics: {metrics_path}")
    print(f"  sample : {sample_path}")
    print(f"  report : {report_path}")
    return 0


if __name__ == "__main__":
    sys.exit(run())
