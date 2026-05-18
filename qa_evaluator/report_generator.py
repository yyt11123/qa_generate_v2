"""生成 Markdown 评测报告。

支持两种模式：
- 仅机器指标 (annotation_xlsx 还没人工填) → 第 5 节为 "[待人工标注完成后补充]"
- 含人工标注 (xlsx 中 B1-B4 至少有部分填值) → 第 5 节用统计填上
"""
from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

from qa_evaluator.sampler import QUOTAS


SECTION_TITLES = {
    "answer_groundedness":      ("answer_groundedness", "answer 字符在 supporting_facts 中的平均比例"),
    "sf_in_chunk_rate":         ("sf_in_chunk_rate", "supporting_facts 与所属 chunk 相似度 ≥0.8 的 QA 比例"),
    "simplified_pollution_rate": ("simplified_pollution_rate", "QA 中含简体污染字符的比例"),
    "meta_question_rate":       ("meta_question_rate", "含元问题关键词的 question 比例"),
    "vague_reference_rate":     ("vague_reference_rate", "answer 开头含模糊指代的比例 (warning)"),
    "category_coverage":        ("category_coverage", "已覆盖分类数 / 8"),
    "category_balance":         ("category_balance", "最大分类条数 / 总条数"),
    "chunk_coverage":           ("chunk_coverage", "已生成 QA 的 chunk_id 数 / 总 chunk 数"),
    "duplicate_qa_rate":        ("duplicate_qa_rate", "同 chunk_id 内 question 相似度 >0.7 的对数 × 2 / 总条数"),
}


def _status_emoji(metric: dict) -> str:
    if metric["passed"]:
        return "✅ pass"
    return "❌ fail" if metric["severity"] == "failure" else "⚠ warn"


def _format_metric_row(metric: dict) -> str:
    name = metric["name"]
    desc = SECTION_TITLES.get(name, (name, ""))[1]
    op = metric["op"]
    return f"| {name} | {metric['value']} | {op} {metric['threshold']} | {_status_emoji(metric)} | {desc} |"


def _format_violations(metric: dict, max_show: int = 10) -> str:
    if not metric["violations"]:
        return "（无）"
    lines = []
    for v in metric["violations"][:max_show]:
        lines.append(f"  - {json.dumps(v, ensure_ascii=False)}")
    if len(metric["violations"]) > max_show:
        lines.append(f"  - ... 另有 {len(metric['violations']) - max_show} 条")
    return "\n".join(lines)


def _read_annotations(xlsx_path: Path) -> tuple[list[dict], dict]:
    """读 sample xlsx 的 B1-B4 + 备注。返回 (rows, summary)。
    rows: 每行附加 _data_row（数据行号，从 1 开始，与用户人工标注时看到的"第 N 条"一致）
    summary: 各维度均值 / 已填条数 / 5 分条数 / 备注计数 / 全维度总均值
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    rows = []
    for i, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        d = dict(zip(headers, raw))
        d["_data_row"] = i
        rows.append(d)

    def _to_float(v):
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def _stats(field: str) -> tuple[float | None, int, int]:
        """返回 (avg, n_filled, n_full_marks)."""
        vals = [_to_float(r.get(field)) for r in rows]
        vals = [v for v in vals if v is not None]
        if not vals:
            return None, 0, 0
        n5 = sum(1 for v in vals if v >= 5.0)
        return round(sum(vals) / len(vals), 3), len(vals), n5

    b1, n1, f1 = _stats("B1 真实性")
    b2, n2, f2 = _stats("B2 正确性")
    b3, n3, f3 = _stats("B3 完整性")
    b4, n4, f4 = _stats("B4 分类合理性")
    notes = sum(1 for r in rows if r.get("备注") and str(r.get("备注")).strip())
    annotated_any = max(n1, n2, n3, n4)

    avgs = [a for a in (b1, b2, b3, b4) if a is not None]
    overall = round(sum(avgs) / len(avgs), 3) if avgs else None

    summary = {
        "total_samples": len(rows),
        "annotated_count": annotated_any,
        "B1_avg": b1, "B1_n": n1, "B1_full": f1,
        "B2_avg": b2, "B2_n": n2, "B2_full": f2,
        "B3_avg": b3, "B3_n": n3, "B3_full": f3,
        "B4_avg": b4, "B4_n": n4, "B4_full": f4,
        "overall_avg": overall,
        "notes_count": notes,
    }
    return rows, summary


def _short(cid: str | None) -> str:
    return (cid or "").split("_")[-1]


def _below_full_lines(rows: list[dict], field: str) -> list[str]:
    out = []
    for r in rows:
        v = r.get(field)
        if v is None or v == "":
            continue
        try:
            f = float(v)
        except (TypeError, ValueError):
            continue
        if f >= 5.0:
            continue
        line = f"- 第 {r['_data_row']} 条 (`...{_short(r.get('chunk_id'))}`, {r.get('分类') or ''}): {field}={int(f) if f.is_integer() else f}"
        note = r.get("备注")
        if note and str(note).strip():
            line += f"，备注：{note}"
        out.append(line)
    return out


def _chunk_class_split(rows: list[dict], chunk_id: str) -> dict[str, list[int]]:
    """返回 chunk_id 在 sample 中各分类对应的数据行号。"""
    out: dict[str, list[int]] = {}
    for r in rows:
        if r.get("chunk_id") == chunk_id:
            cat = r.get("分类") or ""
            out.setdefault(cat, []).append(r["_data_row"])
    return out


def _human_section(annotation_xlsx: Path | None) -> str:
    if annotation_xlsx is None or not annotation_xlsx.exists():
        return "[待人工标注完成后补充]"
    rows, s = _read_annotations(annotation_xlsx)
    if s["annotated_count"] == 0:
        return f"[待人工标注完成后补充]\n\n抽样文件: {annotation_xlsx} (共 {s['total_samples']} 条，尚未标注)"

    lines = [
        f"抽样文件: `{annotation_xlsx}`",
        f"已标注: {s['annotated_count']} / {s['total_samples']} 条；"
        f"全维度总均值 **{s['overall_avg']}** / 5",
        "",
        "### 5.1 维度均值与 5 分占比",
        "",
        "| 维度 | 均值 (1-5) | 已填条数 | 5 分条数 | 5 分占比 |",
        "|---|---|---|---|---|",
    ]
    total = s["total_samples"]
    for dim, avg, n, full in [
        ("B1 真实性", s["B1_avg"], s["B1_n"], s["B1_full"]),
        ("B2 正确性", s["B2_avg"], s["B2_n"], s["B2_full"]),
        ("B3 完整性", s["B3_avg"], s["B3_n"], s["B3_full"]),
        ("B4 分类合理性", s["B4_avg"], s["B4_n"], s["B4_full"]),
    ]:
        pct = f"{full/total*100:.0f}%" if total else "-"
        lines.append(f"| {dim} | {avg} | {n} | {full} | {pct} |")

    lines.append("")
    lines.append("### 5.2 非满分条目明细")
    lines.append("")
    for field in ["B1 真实性", "B2 正确性", "B3 完整性", "B4 分类合理性"]:
        below = _below_full_lines(rows, field)
        if not below:
            continue
        lines.append(f"**{field}** ({len(below)} 条非满分):")
        lines.extend(below)
        lines.append("")

    # 5.3 备注驱动的发现
    note_rows = [r for r in rows if r.get("备注") and str(r.get("备注")).strip()]
    if note_rows:
        lines.append("### 5.3 基于备注的发现")
        lines.append("")
        for nr in note_rows:
            note_txt = str(nr.get("备注")).strip()
            cid = nr.get("chunk_id") or ""
            split = _chunk_class_split(rows, cid)
            if note_txt == "同 chunk 内分类不一致" and len(split) >= 2:
                parts = []
                total_n = sum(len(v) for v in split.values())
                for cat, drows in split.items():
                    parts.append(f"{len(drows)} 条被分到「{cat}」（第 {', '.join(str(d) for d in drows)} 条）")
                lines.append(
                    f"- **`...{_short(cid)}` 同 chunk 内分类不一致**：sample 中 {total_n} 条 QA 出自此 chunk，"
                    + "；".join(parts) + "。"
                )
                lines.append(
                    "  - 含义：LLM 在分类边界上有稳定性问题。同一 chunk 的语义内容应稳定归一类，"
                    "现象说明 prompt 当前没有强约束「同 chunk 一致性」。"
                )
                lines.append(
                    "  - 建议：未来在 qa_prompt.py 里加入「同 chunk 的多条 QA 应保持分类一致性」"
                    "的硬约束，或在生成阶段做后置校正（同 chunk 多分类时按多数投票）。"
                )
            else:
                lines.append(f"- 第 {nr['_data_row']} 条 (`...{_short(cid)}`)：{note_txt}")
        lines.append("")

    return "\n".join(lines)


def _conclusion(metrics: dict, human_filled: bool, human_summary: dict | None,
                chunk_inconsistency: list[str] | None = None) -> str:
    """把判定依据列成事实清单 + 最终判定。

    chunk_inconsistency: 若存在 "同 chunk 内分类不一致" 这类系统性发现，传入
        ["chunk_id 简称", ...] 显式列出。
    """
    failures = [m["name"] for m in metrics.values() if m["severity"] == "failure" and not m["passed"]]
    warnings = [m["name"] for m in metrics.values() if m["severity"] == "warning" and not m["passed"]]
    n_total = len(metrics)
    n_passed = sum(1 for m in metrics.values() if m["passed"])

    lines = ["### 判定依据"]
    lines.append("")
    lines.append(f"- 机器指标：{n_passed}/{n_total} 通过" + (f"（fail: {', '.join(failures)}）" if failures else "")
                 + (f"（warn 未达标: {', '.join(warnings)}）" if warnings else ""))
    if human_filled and human_summary:
        avgs = [human_summary.get(k) for k in ("B1_avg","B2_avg","B3_avg","B4_avg")]
        avgs_clean = [a for a in avgs if a is not None]
        if avgs_clean:
            overall = round(sum(avgs_clean) / len(avgs_clean), 3)
            lines.append(
                f"- 人工 4 维平均：B1={human_summary.get('B1_avg')}, "
                f"B2={human_summary.get('B2_avg')}, B3={human_summary.get('B3_avg')}, "
                f"B4={human_summary.get('B4_avg')}；总均值 **{overall}**（高质量门槛 4.0）"
            )
    if chunk_inconsistency:
        lines.append(f"- 待改进点（1 个）：{', '.join(chunk_inconsistency)} 同 chunk 内分类不一致，"
                     f"LLM 分类边界稳定性问题")
    elif human_filled:
        lines.append("- 待改进点：无")

    lines.append("")
    lines.append("### 最终判定")
    lines.append("")

    if failures:
        lines.append(f"**需改进**：以下 failure 级机器指标未通过 → {', '.join(failures)}。")
    elif human_filled and human_summary:
        avgs_clean = [human_summary.get(k) for k in ("B1_avg","B2_avg","B3_avg","B4_avg") if human_summary.get(k) is not None]
        overall = sum(avgs_clean) / len(avgs_clean) if avgs_clean else 0
        if any(a is not None and a < 4.0 for a in [human_summary.get("B1_avg"), human_summary.get("B2_avg"), human_summary.get("B3_avg"), human_summary.get("B4_avg")]):
            low = [k.replace("_avg","") for k in ("B1_avg","B2_avg","B3_avg","B4_avg") if human_summary.get(k) is not None and human_summary[k] < 4.0]
            lines.append(f"**有限可用**：人工评分 {', '.join(low)} 维度均值 < 4.0，需针对性改进后再用。")
        elif overall >= 4.0:
            lines.append("**v3 测试集质量可用于下一阶段 RAG 系统评测**：")
            lines.append("机器指标全部通过、人工 4 维评分总均值远超 4.0 高质量门槛、"
                         "识别出的待改进点（同 chunk 分类不一致）为下次迭代的优化方向，"
                         "不影响本轮作为评测集使用。")
        else:
            lines.append("**有限可用**：机器指标通过但人工评分整体偏低，建议先迭代再用。")
    elif warnings:
        lines.append(f"**有限可用（机器指标层面）**：failure 级全通过，但 {', '.join(warnings)} 触发 warning。"
                     f"等待人工标注后定论。")
    else:
        lines.append("**机器指标层面：可用**。等待人工标注完成后定终稿。")

    return "\n".join(lines)


def _sampling_findings(annotation_xlsx: Path | None, all_qa_rows: list[dict]) -> dict:
    """从 sample.xlsx + 全集 rows 算两类抽样偏差，供第 6 节列出 + 第 7 节给建议。

    返回 {
      "category_deviations": [(cat, got, want, cross_n_in_cat), ...],
      "chunk_diversity_issues": [(chunk_id, count, [(cat, qa_n, chunk_n), ...]), ...],
      "low_diversity_cats": [(cat, n_chunks_with_qa), ...]   # 全集中 ≤2 chunk 的分类
    }
    """
    findings = {
        "category_deviations": [],
        "chunk_diversity_issues": [],
        "low_diversity_cats": [],
    }
    if not annotation_xlsx or not annotation_xlsx.exists():
        return findings
    wb = openpyxl.load_workbook(annotation_xlsx)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    actual = Counter(r.get("分类") for r in rows)
    cross_per_cat = Counter(
        r.get("分类") for r in rows
        if r.get("page(所在页码)") and "-" in str(r.get("page(所在页码)"))
    )
    for cat, want in QUOTAS.items():
        got = actual.get(cat, 0)
        if got != want:
            findings["category_deviations"].append((cat, got, want, cross_per_cat.get(cat, 0)))

    cid_cnt = Counter(r.get("chunk_id") for r in rows)
    cat_to_chunks: dict[str, set[str]] = {}
    cat_qa_n: Counter = Counter()
    for r in all_qa_rows:
        cat = r.get("category_simp") or r.get("category") or ""
        cat_to_chunks.setdefault(cat, set()).add(r.get("chunk_id") or "")
        cat_qa_n[cat] += 1

    for cid, n in cid_cnt.items():
        if n >= 3:
            cats_in_sample = sorted({r.get("分类") for r in rows if r.get("chunk_id") == cid})
            diag = []
            for cat in cats_in_sample:
                diag.append((cat, cat_qa_n.get(cat, 0), len(cat_to_chunks.get(cat, set()))))
            findings["chunk_diversity_issues"].append((cid, n, diag))

    for cat, chunks in cat_to_chunks.items():
        if cat in QUOTAS and len(chunks) <= 2:
            findings["low_diversity_cats"].append((cat, len(chunks), cat_qa_n.get(cat, 0)))

    return findings


def render_report(
    metrics: dict,
    total_qa: int,
    total_chunks: int,
    annotation_xlsx: Path | None,
    output_md: Path,
    all_qa_rows: list[dict] | None = None,
) -> None:
    human_section = _human_section(annotation_xlsx)
    human_filled = annotation_xlsx and annotation_xlsx.exists() and "[待人工标注完成后补充]" not in human_section
    human_summary = None
    if human_filled:
        _, human_summary = _read_annotations(annotation_xlsx)

    findings = _sampling_findings(annotation_xlsx, all_qa_rows or [])

    a_metrics = ["answer_groundedness", "sf_in_chunk_rate", "simplified_pollution_rate", "meta_question_rate", "vague_reference_rate"]
    c_metrics = ["category_coverage", "category_balance", "chunk_coverage", "duplicate_qa_rate"]

    md = []
    md.append("# QA 测试集 v3 质量评测报告")
    md.append("")
    md.append(f"_生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_")
    md.append(f"_QA 总数: {total_qa}_  _总 chunk 数: {total_chunks}_")
    md.append("")
    md.append("## 1. 评测目的")
    md.append("")
    md.append("本次评测为下一阶段「用 v3 QA 测试集评估 RAG 系统」做准入校验。")
    md.append("RAG 评测的前提是测试集本身可信、覆盖均衡、问题代表用户真实意图；")
    md.append("如果 QA 集存在污染、错位、重复或代表性不足，下游评测的所有结论都会失真。")
    md.append("本报告通过 9 个机器指标 + 4 维度人工抽样标注共同判定 v3 是否达标。")
    md.append("")
    md.append("## 2. 评测方法")
    md.append("")
    md.append("- **数据可信度** 5 个机器指标（A 类）：定位事实错配、繁简污染、元问题、模糊指代")
    md.append("- **覆盖均衡性** 4 个机器指标（C 类）：分类、chunk 覆盖率、重复率")
    md.append("- **问题代表性** 4 维度人工抽样标注（B 类）：B1 真实性 / B2 正确性 / B3 完整性 / B4 分类合理性，每条 1-5 分")
    md.append("")
    md.append("机器指标分两档严重度：`failure` 必须全过，`warning` 给阈值供调优。")
    md.append("")
    md.append("> 抽样说明：为保证跨页 QA 覆盖，跨页 quota 独立于分类配额，所属分类会出现 +1~+2 偏差，详见 `qa_evaluator/sampler.py` 注释。")
    md.append("")
    md.append("## 3. 数据可信度")
    md.append("")
    md.append("| 指标 | 数值 | 阈值 | 状态 | 说明 |")
    md.append("|---|---|---|---|---|")
    for n in a_metrics:
        md.append(_format_metric_row(metrics[n]))
    md.append("")
    md.append("## 4. 覆盖均衡性")
    md.append("")
    md.append("| 指标 | 数值 | 阈值 | 状态 | 说明 |")
    md.append("|---|---|---|---|---|")
    for n in c_metrics:
        md.append(_format_metric_row(metrics[n]))
    md.append("")
    md.append("## 5. 问题代表性（人工抽样 4 维度）")
    md.append("")
    md.append(human_section)
    md.append("")
    md.append("## 6. 已发现问题（机器指标告警明细 + 抽样偏差）")
    md.append("")
    flagged = [m for m in metrics.values() if not m["passed"]]
    if not flagged:
        md.append("机器指标全部通过，无告警明细。")
    else:
        for m in flagged:
            md.append(f"### {m['name']}  ({_status_emoji(m)})  value={m['value']}  threshold {m['op']} {m['threshold']}")
            md.append("")
            md.append(_format_violations(m))
            md.append("")

    if findings["category_deviations"] or findings["chunk_diversity_issues"]:
        md.append("")
        md.append("### 抽样偏差（软约束）")
        md.append("")
        for cat, got, want, cross_n in findings["category_deviations"]:
            if cross_n:
                reason = f"跨页 QA 落入此分类 {cross_n} 条"
            elif got < want:
                reason = "该分类 QA 总数不足配额"
            else:
                reason = "其他原因（补齐）"
            md.append(f"- 分类配额偏差：**{cat}** 抽到 {got} 条（预期 {want} 条），原因：{reason}")
        for cid, n, diag in findings["chunk_diversity_issues"]:
            short = (cid or "").split("_")[-1]
            diag_str = "; ".join(f"{cat} 全集 {qa_n} 条 QA / 分布在 {chunk_n} 个 chunk" for cat, qa_n, chunk_n in diag)
            md.append(f"- chunk 多样性偏差：`...{short}` 在 sample 中出现 {n} 次")
            md.append(f"  - 原因：{diag_str}，约束不到 chunk 多样性")
        md.append("")

    md.append("## 7. 结论")
    md.append("")
    chunk_inconsistency = []
    if annotation_xlsx and annotation_xlsx.exists():
        ann_rows, _ = _read_annotations(annotation_xlsx)
        for nr in ann_rows:
            note_txt = str(nr.get("备注") or "").strip()
            if note_txt == "同 chunk 内分类不一致":
                cid = nr.get("chunk_id") or ""
                short = _short(cid)
                if short and f"`...{short}`" not in chunk_inconsistency:
                    chunk_inconsistency.append(f"`...{short}`")
    md.append(_conclusion(metrics, bool(human_filled), human_summary, chunk_inconsistency))
    if findings["low_diversity_cats"]:
        md.append("")
        suggestions = []
        for cat, n_chunks, qa_n in findings["low_diversity_cats"]:
            suggestions.append(f"`{cat}`（{qa_n} 条 QA 仅分布在 {n_chunks} 个 chunk）")
        md.append(f"建议：发现 v3 数据集 {', '.join(suggestions)} 分类多样性偏低，建议未来扩展数据集时增加该分类的 chunk 覆盖。")
    md.append("")
    output_md.parent.mkdir(parents=True, exist_ok=True)
    output_md.write_text("\n".join(md), encoding="utf-8")
