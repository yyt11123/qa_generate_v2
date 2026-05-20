"""RAG 置信度评测 CLI 入口。

用法：
  # 仅解析日志（无 API 调用，安全）
  python -m qa_evaluator.confidence.confidence_runner --parse-only

  # 跑完整 faithfulness 评估（有 API 调用）
  python -m qa_evaluator.confidence.confidence_runner

输出：
  output/rag_logs_parsed.json       —— 解析后的结构化日志
  output/rag_faithfulness.json      —— 每条 QA 的 statement 级判定
  output/rag_confidence_summary.xlsx —— 汇总表（Q 维度）
"""
from __future__ import annotations

import argparse
import json
import logging
from pathlib import Path

from qa_evaluator.confidence.log_parser import parse_log, write_parsed

logger = logging.getLogger(__name__)

DEFAULT_LOG = Path("inputs/full_retrieval_batch_test_output.txt")
DEFAULT_PARSED = Path("output/rag_logs_parsed.json")
DEFAULT_RESULT = Path("output/confidence_per_question.json")
DEFAULT_SUMMARY_XLSX = Path("output/confidence_summary.xlsx")


def _print_parse_summary(parsed: list[dict]) -> None:
    print(f"Parsed {len(parsed)} questions")
    for q in parsed:
        print(
            f"  Q{q['question_id']}: chunks={len(q['retrieved_chunks'])}, "
            f"answer_chars={len(q['answer'])}"
        )


def _write_summary_xlsx(result: dict, out_path: Path) -> None:
    """单页 xlsx：每行一条 QA 的总分 + 各等级数量。带表头 / 列宽 / 冻结样式。"""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = []
    for r in result["per_question"]:
        rows.append({
            "question_id": r["question_id"],
            "question": r["question"],
            "n_statements": r["n_statements"],
            "n_yes": r["n_yes"],
            "n_partial": r["n_partial"],
            "n_no": r["n_no"],
            "score": r["score"],
            "level": r["level"],
            "llm_calls": r["llm_calls"],
        })
    summary = result["summary"]
    rows.append({
        "question_id": "AVG",
        "question": "(汇总)",
        "n_statements": sum(r["n_statements"] for r in result["per_question"]),
        "n_yes": sum(r["n_yes"] for r in result["per_question"]),
        "n_partial": sum(r["n_partial"] for r in result["per_question"]),
        "n_no": sum(r["n_no"] for r in result["per_question"]),
        "score": summary["avg_score"],
        "level": summary["avg_level"],
        "llm_calls": summary["total_llm_calls"],
    })
    df = pd.DataFrame(rows)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(out_path, index=False, engine="openpyxl")

    widths = {
        "question_id": 12, "question": 50, "n_statements": 12,
        "n_yes": 8, "n_partial": 10, "n_no": 8,
        "score": 8, "level": 10, "llm_calls": 10,
    }
    wb = load_workbook(out_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for idx, h in enumerate(headers, start=1):
        if h in widths:
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths[h]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 28
    body_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = body_align
    ws.freeze_panes = "A2"
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="RAG faithfulness 置信度评测")
    parser.add_argument(
        "--log", type=Path, default=DEFAULT_LOG, help="原始 RAG 日志路径",
    )
    parser.add_argument(
        "--parsed", type=Path, default=DEFAULT_PARSED, help="解析输出 JSON 路径",
    )
    parser.add_argument(
        "--result", type=Path, default=DEFAULT_RESULT, help="faithfulness 结果 JSON 路径",
    )
    parser.add_argument(
        "--summary-xlsx", type=Path, default=DEFAULT_SUMMARY_XLSX, help="汇总 xlsx 路径",
    )
    parser.add_argument(
        "--parse-only", action="store_true",
        help="只解析日志，不调 LLM；产出 rag_logs_parsed.json 后退出",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")

    parsed = parse_log(args.log)
    write_parsed(parsed, args.parsed)
    _print_parse_summary(parsed)
    print(f"Wrote {args.parsed}")

    if args.parse_only:
        print("--parse-only set, stopping before LLM calls.")
        return

    # 仅在用户明确要求跑完整评估时导入 faithfulness（避免触发 client 初始化）
    from qa_evaluator.confidence.faithfulness import evaluate_all

    result = evaluate_all(parsed)
    args.result.parent.mkdir(parents=True, exist_ok=True)
    args.result.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {args.result}")

    _write_summary_xlsx(result, args.summary_xlsx)
    print(f"Wrote {args.summary_xlsx}")

    summary = result["summary"]
    print()
    print(f"=== Summary ===")
    print(f"  questions:      {summary['n_questions']}")
    print(f"  avg_score:      {summary['avg_score']} ({summary['avg_level']})")
    print(f"  total LLM calls: {summary['total_llm_calls']}")


if __name__ == "__main__":
    main()
